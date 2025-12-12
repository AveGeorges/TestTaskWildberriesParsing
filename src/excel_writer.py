import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


COLUMNS = [
    ("Ссылка на товар", "url", 50),
    ("Артикул", "article", 15),
    ("Название", "name", 40),
    ("Бренд", "brand", 20),
    ("Цена (руб.)", "price_rub", 15),
    ("Описание", "description_clean", 60),
    ("Изображения", "images_str", 80),
    ("Характеристики", "characteristics_str", 60),
    ("Продавец", "seller_name", 25),
    ("Ссылка на продавца", "seller_url", 50),
    ("Размеры", "sizes_str", 30),
    ("Остатки", "stock", 12),
    ("Рейтинг", "rating", 10),
    ("Отзывы", "feedbacks_count", 12),
    ("Страна", "country", 20),
]


def _style_header(ws):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    col_num = 1
    for col_info in COLUMNS:
        name = col_info[0]
        width = col_info[2]
        cell = ws.cell(row=1, column=col_num, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_num)].width = width
        col_num += 1


def _write_row(ws, row, product):
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    for i, (_, attr, _) in enumerate(COLUMNS, 1):
        val = getattr(product, attr, "")
        cell = ws.cell(row=row, column=i, value=val)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border


def save_xlsx(products, filepath, sheet_name="Товары"):
    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    _style_header(ws)
    
    for i, p in enumerate(products, 2):
        _write_row(ws, i, p)
    
    ws.freeze_panes = "A2"
    wb.save(filepath)
    
    logger.info(f"Сохранено {len(products)} товаров в {filepath}")
    return filepath


def save_filtered(products, filepath, filter_func):
    filtered = [p for p in products if filter_func(p)]
    save_xlsx(filtered, filepath, "Отфильтрованные")
    return len(filtered)

export_to_xlsx = save_xlsx
export_filtered = lambda products, path, func, **kw: (Path(path), save_filtered(products, path, func))
